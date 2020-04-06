import pygame
import sys
import math
from settings import *
from player_class import *
from enemy_class import *
from grid_class import *
from openpyxl import Workbook
from random import *
pygame.init()
vec = pygame.math.Vector2


class App:
    def __init__(self):
        self.screen = pygame.display.set_mode((WIDTH, HEIGHT))
        self.clock = pygame.time.Clock()
        self.running = True
        self.state = 'playing'
        self.cell_width = MAZE_WIDTH // COLS
        self.cell_height = MAZE_HEIGHT // ROWS
        self.walls = []
        self.coins = []
        self.enemies = []
        self.grid_tiles = []
        self.e_pos = []
        self.g_pos = []
        self.p_pos = None
        self.load()
        self.player = Player(self, self.p_pos)
        self.make_enemies()
        self.make_grid()
        self.value = int()

        # Make a data set & column titles
        self.wb = Workbook()
        self.ws = self.wb.active
        self.cell_row = 1
        self.ws.cell(row=self.cell_row, column=1, value="GRID VALUES")
        self.ws.cell(row=self.cell_row, column=26, value="PLAYER DIRECTION")
        self.ws.cell(row=self.cell_row, column=27, value="PLAYER SCORE")

    def run(self):
        while self.running:
            if self.state == 'playing':
                self.playing_events()
                self.playing_update()
                self.playing_draw()
            else:
                self.running = False
            self.clock.tick(FPS)
        pygame.quit()
        sys.exit()

#################################################### HELPER FUNCTIONS ##################################################

    def draw_text(self, words, screen, pos, size, colour, font_name, centered=False):
        font = pygame.font.SysFont(font_name, size)
        text = font.render(words, False, colour)
        text_size = text.get_size()
        if centered:
            pos[0] = pos[0] - text_size[0] // 2
            pos[1] = pos[1] - text_size[1] // 2
        screen.blit(text, pos)

    def load(self):
        self.background = pygame.image.load('maze.png')
        self.background = pygame.transform.scale(self.background, (MAZE_WIDTH, MAZE_HEIGHT))

        # Opening walls file
        # Creating walls list with co-ords of walls
        # Stored as a vector
        with open("walls", 'r') as file:
            for yidx, line in enumerate(file):
                for xidx, char in enumerate(line):
                    if char == "1":
                        self.walls.append(vec(xidx, yidx))
                    elif char == "C":
                        self.coins.append(vec(xidx, yidx))
                        self.g_pos.append(vec(xidx, yidx))
                    elif char == "P":
                        self.p_pos = vec(xidx, yidx)
                    elif char in ["2", "3", "4", "5"]:
                        self.e_pos.append(vec(xidx, yidx))
                    elif char == "B":
                        pygame.draw.rect(self.background, BLACK, (xidx * self.cell_width, yidx * self.cell_height,
                                                                  self.cell_width, self.cell_height))

    def make_enemies(self):
        for idx, pos in enumerate(self.e_pos):
            self.enemies.append(Enemy(self, pos, idx))

    def make_grid(self):
        for idx, pos in enumerate(self.g_pos):
            self.grid_tiles.append(Grid(self, pos, idx))

    def draw_grid(self):
        for x in range(WIDTH // self.cell_width):
            pygame.draw.line(self.background, GREY, (x * self.cell_width, 0), (x * self.cell_width, HEIGHT))
        for x in range(HEIGHT // self.cell_height):
            pygame.draw.line(self.background, GREY, (0, x * self.cell_height), (WIDTH, x * self.cell_height))



################################################### INTRO FUNCTIONS ####################################################

    def start_events(self):
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                self.running = False

    def start_update(self):
        pass

    def start_draw(self):
        self.screen.fill(BLACK)
        self.draw_text('PUSH SPACE BAR', self.screen, [WIDTH // 2, HEIGHT // 2 - 50], START_TEXT_SIZE, (170, 132, 58),
                       START_FONT, centered=True)
        self.draw_text('1 PLAYER ONLY', self.screen, [WIDTH // 2, HEIGHT // 2 + 50], START_TEXT_SIZE, (44, 167, 198),
                       START_FONT, centered=True)
        self.draw_text('HIGH SCORE', self.screen, [4, 0], START_TEXT_SIZE, (255, 255, 255), START_FONT)
        pygame.display.update()

################################################## PLAYING FUNCTIONS ###################################################

    def playing_events(self):
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                self.running = False

    def playing_update(self):
        self.player.update()
        for enemy in self.enemies:
            enemy.update()
        for tiles in self.grid_tiles:
            tiles.update()

    def playing_draw(self):
        self.screen.fill(BLACK)
        self.screen.blit(self.background, (TOP_BOTTOM_BUFFER//2, TOP_BOTTOM_BUFFER//2))
        self.draw_coins()
        # self.draw_grid()
        self.draw_text('CURRENT SCORE: {}'.format(self.player.current_score), self.screen, [60, 0], 18, WHITE,
                       START_FONT)
        self.draw_text('HIGH SCORE: 0', self.screen, [WIDTH//2+60, 0], 18, WHITE, START_FONT)
        self.player.draw()
        for enemy in self.enemies:
            enemy.draw()
        for tiles in self.grid_tiles:
            tiles.draw()
        pygame.display.update()

    def draw_coins(self):
        for coin in self.coins:
            pygame.draw.circle(self.screen, WHITE, (int(coin.x * self.cell_width) + self.cell_width//2 + TOP_BOTTOM_BUFFER
                                                  // 2, int(coin.y * self.cell_height) + self.cell_height//2 +
                                                  TOP_BOTTOM_BUFFER//2), 5)
